import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import random
from datetime import datetime, timedelta

random.seed(42)
np.random.seed(42)

# ── RAW DATA GENERATION ──────────────────────────────────────────────────────
regions   = ["North", "South", "East", "West", "Central"]
products  = ["Laptop", "Phone", "Tablet", "Monitor", "Keyboard", "Mouse",
             "Headset", "Webcam"]
customers = [f"Customer_{i:03d}" for i in range(1, 51)]

def rand_date(start="2024-01-01", end="2024-12-31"):
    s = datetime.strptime(start, "%Y-%m-%d")
    e = datetime.strptime(end, "%Y-%m-%d")
    return s + timedelta(days=random.randint(0, (e - s).days))

prices = {"Laptop":999,"Phone":699,"Tablet":499,"Monitor":399,
          "Keyboard":149,"Mouse":79,"Headset":199,"Webcam":129}

n = 500
orders = []
order_ids = set()
for i in range(n):
    oid = f"ORD-{random.randint(1000,9999)}"
    orders.append({
        "Order_ID":   oid,
        "Order_Date": rand_date(),
        "Customer_ID": random.choice(customers),
        "Product":    random.choice(products),
        "Region":     random.choice(regions),
        "Quantity":   random.randint(1, 10),
        "Unit_Price": None,   # filled below
        "Discount":   round(random.uniform(0, 0.3), 2),
        "Status":     random.choice(["Completed","Pending","Cancelled"])
    })
    order_ids.add(oid)

for o in orders:
    o["Unit_Price"] = prices[o["Product"]] * random.uniform(0.9, 1.1)
    o["Unit_Price"] = round(o["Unit_Price"], 2)

df = pd.DataFrame(orders)

# ── INJECT DATA QUALITY ISSUES ───────────────────────────────────────────────
# Missing values
for idx in random.sample(range(n), 20):
    df.at[idx, "Customer_ID"] = None
for idx in random.sample(range(n), 15):
    df.at[idx, "Unit_Price"] = None
for idx in random.sample(range(n), 10):
    df.at[idx, "Region"] = None

# Invalid orders
for idx in random.sample(range(n), 12):
    df.at[idx, "Quantity"] = random.choice([-1, -5, 0])
for idx in random.sample(range(n), 8):
    df.at[idx, "Unit_Price"] = random.choice([-99, -1])

# Duplicate transactions (exact duplicates)
dupes = df.sample(15).copy()
df = pd.concat([df, dupes], ignore_index=True)

df["Revenue"] = df["Quantity"] * df["Unit_Price"] * (1 - df["Discount"])
df["Month"]   = pd.to_datetime(df["Order_Date"]).dt.strftime("%Y-%m")

# ── WORKBOOK ─────────────────────────────────────────────────────────────────
wb = Workbook()

# color palette
C_DARK   = "1E2A3A"
C_BLUE   = "2563EB"
C_GREEN  = "10B981"
C_RED    = "EF4444"
C_AMBER  = "F59E0B"
C_LIGHT  = "F1F5F9"
C_WHITE  = "FFFFFF"
C_HEADER = "1E40AF"
C_PASS   = "D1FAE5"
C_FAIL   = "FEE2E2"
C_WARN   = "FEF3C7"

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def write_header_row(ws, row, cols, bg=C_HEADER, fg=C_WHITE, bold=True):
    for c, txt in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=txt)
        cell.font      = Font(bold=bold, color=fg, name="Arial", size=10)
        cell.fill      = hdr_fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()

def style_data_cell(cell, align="center"):
    cell.font      = Font(name="Arial", size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = thin_border()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – RAW DATA
# ══════════════════════════════════════════════════════════════════════════════
ws_raw = wb.active
ws_raw.title = "Raw_Data"
ws_raw.sheet_view.showGridLines = False
ws_raw.freeze_panes = "A2"

cols = ["Order_ID","Order_Date","Customer_ID","Product","Region",
        "Quantity","Unit_Price","Discount","Status","Revenue","Month"]
write_header_row(ws_raw, 1, cols)

for r, row in enumerate(df[cols].itertuples(index=False), 2):
    for c, val in enumerate(row, 1):
        cell = ws_raw.cell(row=r, column=c, value=val)
        style_data_cell(cell, "left" if c in (1,3,4,5,9) else "center")
        if isinstance(val, float):
            if c in (7, 10):  # price / revenue
                cell.number_format = '#,##0.00'
            elif c == 8:
                cell.number_format = '0%'

ws_raw.row_dimensions[1].height = 30
widths = [14,13,14,11,10,9,11,10,12,12,10]
for i, w in enumerate(widths, 1):
    ws_raw.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – DATA QUALITY REPORT
# ══════════════════════════════════════════════════════════════════════════════
ws_dq = wb.create_sheet("Data_Quality")
ws_dq.sheet_view.showGridLines = False

def banner(ws, row, text, bg=C_DARK):
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color=C_WHITE, size=12, name="Arial")
    c.fill      = hdr_fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = thin_border()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.row_dimensions[row].height = 28

# ── Section A: Summary KPIs ───────────────────────────────────────────────────
banner(ws_dq, 1, "📊  DATA QUALITY EXECUTIVE SUMMARY")
ws_dq.row_dimensions[1].height = 32

kpi_labels = ["Total Records","Missing Values","Invalid Orders",
              "Duplicate Rows","Clean Records","Quality Score"]

total     = len(df)
missing   = int(df[["Customer_ID","Unit_Price","Region"]].isnull().any(axis=1).sum())
invalid   = int(((df["Quantity"] <= 0) | (df["Unit_Price"] < 0)).sum())
dupl      = int(df.duplicated(subset=["Order_ID","Customer_ID","Product",
                                       "Quantity","Unit_Price"]).sum())
clean     = total - missing - invalid - dupl
score     = round(clean / total * 100, 1)

kpi_vals  = [total, missing, invalid, dupl, clean, f"{score}%"]
kpi_colors= [C_BLUE, C_RED, C_AMBER, C_AMBER, C_GREEN,
             C_GREEN if score >= 80 else C_RED]

write_header_row(ws_dq, 3, kpi_labels, bg=C_HEADER)
ws_dq.row_dimensions[3].height = 24

for c, (val, col) in enumerate(zip(kpi_vals, kpi_colors), 1):
    cell = ws_dq.cell(row=4, column=c, value=val)
    cell.font      = Font(bold=True, size=14, color=C_WHITE, name="Arial")
    cell.fill      = hdr_fill(col)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border()
    ws_dq.row_dimensions[4].height = 36

# ── Section B: Missing Values Detail ─────────────────────────────────────────
banner(ws_dq, 6, "❌  MISSING VALUES ANALYSIS", bg="991B1B")
cols_check = ["Customer_ID", "Unit_Price", "Region"]
mv_headers = ["Column", "Missing Count", "Missing %", "Impact", "Action Required"]
write_header_row(ws_dq, 7, mv_headers, bg="DC2626")
ws_dq.row_dimensions[7].height = 22

impacts = {"Customer_ID":"Revenue attribution lost",
           "Unit_Price":"Revenue calc impossible",
           "Region":"Regional analysis skewed"}
actions = {"Customer_ID":"Lookup from CRM / flag for review",
           "Unit_Price":"Impute median or exclude",
           "Region":"Geo-code from customer address"}

for r, col in enumerate(cols_check, 8):
    cnt = int(df[col].isnull().sum())
    pct = f"{cnt/total*100:.1f}%"
    for c, val in enumerate([col, cnt, pct, impacts[col], actions[col]], 1):
        cell = ws_dq.cell(row=r, column=c, value=val)
        cell.fill      = hdr_fill("FEE2E2" if cnt > 0 else C_PASS)
        style_data_cell(cell, "left" if c >= 4 else "center")

# ── Section C: Invalid Orders ─────────────────────────────────────────────────
banner(ws_dq, 12, "⚠️  INVALID ORDERS DETECTED", bg="92400E")
inv_heads = ["Rule","Count","% of Total","Severity","Recommended Action"]
write_header_row(ws_dq, 13, inv_heads, bg="D97706")
ws_dq.row_dimensions[13].height = 22

neg_qty  = int((df["Quantity"] <= 0).sum())
neg_price= int((df["Unit_Price"] < 0).sum())
zero_qty = int((df["Quantity"] == 0).sum())

rules = [
    ("Negative Quantity",     neg_qty,    "HIGH",   "Remove or reclassify as returns"),
    ("Zero Quantity",         zero_qty,   "MEDIUM", "Verify if sample/gift; remove if error"),
    ("Negative Unit Price",   neg_price,  "HIGH",   "Correct pricing data or remove"),
]
sev_colors = {"HIGH": C_FAIL, "MEDIUM": C_WARN, "LOW": C_PASS}

for r, (rule, cnt, sev, action) in enumerate(rules, 14):
    pct = f"{cnt/total*100:.1f}%"
    for c, val in enumerate([rule, cnt, pct, sev, action], 1):
        cell = ws_dq.cell(row=r, column=c, value=val)
        cell.fill = hdr_fill(sev_colors[sev])
        style_data_cell(cell, "left" if c in (1,4,5) else "center")
        if c == 4:
            cell.font = Font(bold=True, name="Arial", size=9,
                             color="991B1B" if sev=="HIGH" else "92400E")

# ── Section D: Duplicates ──────────────────────────────────────────────────────
banner(ws_dq, 18, "🔁  DUPLICATE TRANSACTION ANALYSIS", bg="1E3A5F")
dup_heads = ["Check Type","Duplicates Found","% of Total","Action"]
write_header_row(ws_dq, 19, dup_heads, bg=C_BLUE)
ws_dq.row_dimensions[19].height = 22

dup_exact = int(df.duplicated(subset=["Order_ID","Customer_ID","Product",
                                       "Quantity","Unit_Price"]).sum())
dup_oid   = int(df.duplicated(subset=["Order_ID"]).sum())

for r, (chk, cnt, act) in enumerate([
    ("Exact Row Duplicates",    dup_exact, "Remove duplicates; keep first occurrence"),
    ("Duplicate Order IDs",     dup_oid,   "Investigate – possible double-billing"),
], 20):
    pct = f"{cnt/total*100:.1f}%"
    for c, val in enumerate([chk, cnt, pct, act], 1):
        cell = ws_dq.cell(row=r, column=c, value=val)
        cell.fill = hdr_fill(C_WARN if cnt > 0 else C_PASS)
        style_data_cell(cell, "left" if c in (1,4) else "center")

# ── Section E: Cleaned data stats ────────────────────────────────────────────
banner(ws_dq, 23, "✅  CLEAN DATASET SUMMARY (POST-QUALITY FILTER)", bg="065F46")
cl_heads = ["Metric","Value"]
write_header_row(ws_dq, 24, cl_heads, bg=C_GREEN)

df_clean = df.copy()
df_clean = df_clean.dropna(subset=["Customer_ID","Unit_Price","Region"])
df_clean = df_clean[(df_clean["Quantity"] > 0) & (df_clean["Unit_Price"] > 0)]
df_clean = df_clean.drop_duplicates(subset=["Order_ID","Customer_ID","Product",
                                             "Quantity","Unit_Price"])
df_clean["Revenue"] = df_clean["Quantity"] * df_clean["Unit_Price"] * (1 - df_clean["Discount"])

clean_stats = [
    ("Clean Record Count",  len(df_clean)),
    ("Total Revenue (Clean)", f"${df_clean['Revenue'].sum():,.0f}"),
    ("Avg Order Value",     f"${df_clean['Revenue'].mean():,.0f}"),
    ("Unique Customers",    df_clean["Customer_ID"].nunique()),
    ("Date Range",          f"{df_clean['Order_Date'].min().date()} → {df_clean['Order_Date'].max().date()}"),
]
for r, (m, v) in enumerate(clean_stats, 25):
    c1 = ws_dq.cell(row=r, column=1, value=m)
    c2 = ws_dq.cell(row=r, column=2, value=v)
    c1.fill = hdr_fill(C_PASS); c2.fill = hdr_fill(C_PASS)
    style_data_cell(c1, "left"); style_data_cell(c2, "left")
    c1.font = Font(bold=True, name="Arial", size=9)

# column widths for DQ sheet
dq_widths = [32, 16, 12, 22, 42]
for i, w in enumerate(dq_widths, 1):
    ws_dq.column_dimensions[get_column_letter(i)].width = w
for i in range(6, 9):
    ws_dq.column_dimensions[get_column_letter(i)].width = 5

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – SALES DASHBOARD (clean data)
# ══════════════════════════════════════════════════════════════════════════════
ws_dash = wb.create_sheet("Sales_Dashboard")
ws_dash.sheet_view.showGridLines = False

# ── Title Banner ──────────────────────────────────────────────────────────────
ws_dash.merge_cells("A1:N1")
title = ws_dash["A1"]
title.value     = "📈  SALES ANALYTICS DASHBOARD  |  FY 2024  |  Clean Data Only"
title.font      = Font(bold=True, color=C_WHITE, size=14, name="Arial")
title.fill      = hdr_fill(C_DARK)
title.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[1].height = 36

# ── KPI Row ───────────────────────────────────────────────────────────────────
kpi_data = [
    ("Total Revenue",    f"${df_clean['Revenue'].sum():,.0f}",    C_BLUE),
    ("Total Orders",     f"{len(df_clean):,}",                     C_GREEN),
    ("Avg Order Value",  f"${df_clean['Revenue'].mean():,.0f}",   "7C3AED"),
    ("Top Region",       df_clean.groupby('Region')['Revenue'].sum().idxmax(), C_AMBER),
]
ws_dash.row_dimensions[3].height = 18
ws_dash.row_dimensions[4].height = 34
ws_dash.row_dimensions[5].height = 18

for i, (label, val, color) in enumerate(kpi_data):
    col_start = i * 3 + 1
    col_end   = col_start + 2
    # label
    ws_dash.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_end)
    lc = ws_dash.cell(row=3, column=col_start, value=label)
    lc.font      = Font(bold=True, color=color, size=9, name="Arial")
    lc.fill      = hdr_fill(C_LIGHT)
    lc.alignment = Alignment(horizontal="center")
    # value
    ws_dash.merge_cells(start_row=4, start_column=col_start, end_row=4, end_column=col_end)
    vc = ws_dash.cell(row=4, column=col_start, value=val)
    vc.font      = Font(bold=True, color=color, size=16, name="Arial")
    vc.fill      = hdr_fill(C_LIGHT)
    vc.alignment = Alignment(horizontal="center", vertical="center")

# ── Monthly Revenue Table ─────────────────────────────────────────────────────
monthly = df_clean.groupby("Month")["Revenue"].sum().reset_index().sort_values("Month")

banner(ws_dash, 7, "Monthly Revenue Trend", bg=C_BLUE)
write_header_row(ws_dash, 8, ["Month", "Revenue ($)"], bg=C_HEADER)
for r, row in enumerate(monthly.itertuples(index=False), 9):
    c1 = ws_dash.cell(row=r, column=1, value=row.Month)
    c2 = ws_dash.cell(row=r, column=2, value=round(row.Revenue, 0))
    c1.number_format = "@"
    c2.number_format = "#,##0"
    style_data_cell(c1); style_data_cell(c2)
    c1.fill = hdr_fill(C_LIGHT if r % 2 == 0 else C_WHITE)
    c2.fill = hdr_fill(C_LIGHT if r % 2 == 0 else C_WHITE)

# Line chart – monthly revenue
chart_line = LineChart()
chart_line.title  = "Revenue Trend 2024"
chart_line.style  = 10
chart_line.y_axis.title = "Revenue ($)"
chart_line.x_axis.title = "Month"
chart_line.width  = 18
chart_line.height = 10

data_ref  = Reference(ws_dash, min_col=2, min_row=8,
                       max_row=8 + len(monthly))
cats_ref  = Reference(ws_dash, min_col=1, min_row=9,
                       max_row=8 + len(monthly))
chart_line.add_data(data_ref, titles_from_data=True)
chart_line.set_categories(cats_ref)
chart_line.series[0].graphicalProperties.line.solidFill = C_BLUE
chart_line.series[0].graphicalProperties.line.width     = 25000
ws_dash.add_chart(chart_line, "D7")

# ── Top Customers Table ───────────────────────────────────────────────────────
top_cust = (df_clean.groupby("Customer_ID")["Revenue"]
            .sum().nlargest(10).reset_index())
top_cust.columns = ["Customer", "Revenue"]

row_start = 22
banner(ws_dash, row_start, "Top 10 Customers by Revenue", bg="065F46")
write_header_row(ws_dash, row_start+1, ["#","Customer","Revenue ($)","Share"], bg=C_GREEN)

total_rev = df_clean["Revenue"].sum()
for r, row in enumerate(top_cust.itertuples(index=False), row_start+2):
    share = row.Revenue / total_rev * 100
    vals  = [r - row_start - 1, row.Customer, round(row.Revenue, 0), f"{share:.1f}%"]
    for c, val in enumerate(vals, 1):
        cell = ws_dash.cell(row=r, column=c, value=val)
        cell.fill = hdr_fill(C_LIGHT if r % 2 == 0 else C_WHITE)
        style_data_cell(cell, "left" if c == 2 else "center")
        if c == 3:
            cell.number_format = "#,##0"

# ── Region Performance Table ──────────────────────────────────────────────────
region_perf = (df_clean.groupby("Region")
               .agg(Revenue=("Revenue","sum"),
                    Orders=("Order_ID","count"),
                    AvgOrder=("Revenue","mean"))
               .reset_index().sort_values("Revenue", ascending=False))

row_start2 = 22
banner(ws_dash, row_start2, "Region Performance", bg="1E3A5F")
write_header_row(ws_dash, row_start2+1,
                 ["Region","Revenue ($)","Orders","Avg Order ($)","Revenue Share"],
                 bg=C_HEADER)

for r, row in enumerate(region_perf.itertuples(index=False), row_start2+2):
    share = row.Revenue / total_rev * 100
    vals  = [row.Region, round(row.Revenue,0), row.Orders,
             round(row.AvgOrder,0), f"{share:.1f}%"]
    for c, val in enumerate(vals, 1):
        cell = ws_dash.cell(row=r, column=c, value=val)
        cell.fill = hdr_fill(C_LIGHT if r % 2 == 0 else C_WHITE)
        style_data_cell(cell, "left" if c == 1 else "center")
        if c in (2,4):
            cell.number_format = "#,##0"

# Bar chart – region revenue
chart_bar = BarChart()
chart_bar.type   = "col"
chart_bar.title  = "Revenue by Region"
chart_bar.style  = 10
chart_bar.y_axis.title = "Revenue ($)"
chart_bar.width  = 14
chart_bar.height = 10
chart_bar.grouping = "clustered"

b_data = Reference(ws_dash, min_col=9, min_row=row_start2+1,
                   max_row=row_start2+1+len(region_perf))
b_cats = Reference(ws_dash, min_col=8, min_row=row_start2+2,
                   max_row=row_start2+1+len(region_perf))
chart_bar.add_data(b_data, titles_from_data=True)
chart_bar.set_categories(b_cats)
chart_bar.series[0].graphicalProperties.solidFill = C_BLUE
ws_dash.add_chart(chart_bar, "D22")

# ── Product Performance Table ─────────────────────────────────────────────────
prod_perf = (df_clean.groupby("Product")
             .agg(Revenue=("Revenue","sum"), Units=("Quantity","sum"))
             .reset_index().sort_values("Revenue", ascending=False))

row_start3 = 35
banner(ws_dash, row_start3, "Product Performance", bg="4C1D95")
write_header_row(ws_dash, row_start3+1,
                 ["Product","Revenue ($)","Units Sold","Revenue Share"],
                 bg="7C3AED")

for r, row in enumerate(prod_perf.itertuples(index=False), row_start3+2):
    share = row.Revenue / total_rev * 100
    vals  = [row.Product, round(row.Revenue,0), row.Units, f"{share:.1f}%"]
    for c, val in enumerate(vals, 1):
        cell = ws_dash.cell(row=r, column=c, value=val)
        cell.fill = hdr_fill(C_LIGHT if r % 2 == 0 else C_WHITE)
        style_data_cell(cell, "left" if c == 1 else "center")
        if c == 2:
            cell.number_format = "#,##0"

# Column widths – dashboard
dash_widths = {1:4, 2:16, 3:14, 4:14, 5:14, 6:14, 7:5,
               8:14, 9:14, 10:12, 11:12, 12:14}
for col, w in dash_widths.items():
    ws_dash.column_dimensions[get_column_letter(col)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – CLEAN DATA (for Power BI)
# ══════════════════════════════════════════════════════════════════════════════
ws_clean = wb.create_sheet("Clean_Data_PowerBI")
ws_clean.sheet_view.showGridLines = False
ws_clean.freeze_panes = "A2"

clean_cols = ["Order_ID","Order_Date","Customer_ID","Product","Region",
              "Quantity","Unit_Price","Discount","Status","Revenue","Month"]
write_header_row(ws_clean, 1, clean_cols)

for r, row in enumerate(df_clean[clean_cols].itertuples(index=False), 2):
    for c, val in enumerate(row, 1):
        cell = ws_clean.cell(row=r, column=c, value=val)
        style_data_cell(cell, "left" if c in (1,3,4,5,9) else "center")
        if isinstance(val, float):
            if c in (7, 10): cell.number_format = '#,##0.00'
            elif c == 8:     cell.number_format = '0%'

ws_clean.row_dimensions[1].height = 30
for i, w in enumerate(widths, 1):
    ws_clean.column_dimensions[get_column_letter(i)].width = w

# ── Tab colors ────────────────────────────────────────────────────────────────
ws_raw.sheet_properties.tabColor   = "64748B"
ws_dq.sheet_properties.tabColor    = "EF4444"
ws_dash.sheet_properties.tabColor  = C_BLUE
ws_clean.sheet_properties.tabColor = C_GREEN

# ── Sheet order ───────────────────────────────────────────────────────────────
wb.active = wb["Sales_Dashboard"]

out = "/mnt/user-data/outputs/Sales_Analytics_Dashboard.xlsx"
wb.save(out)
print("Saved:", out)
print(f"Raw rows: {len(df)} | Clean rows: {len(df_clean)} | Quality: {score}%")
